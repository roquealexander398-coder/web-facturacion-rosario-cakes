from django.shortcuts import render, get_object_or_404, redirect
from django.urls import reverse
from django.http import HttpResponse, JsonResponse
from django.db.models import Sum, Max
from django.contrib.auth.decorators import login_required, permission_required
from django.db import transaction
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from decimal import Decimal
import json
import openpyxl

from .models import Sale, SaleDetail
from ..products.models import Product, InventoryMovement
from ..clients.models import Client
from ..audit.models import AuditLog
from ..company.models import CompanyConfig
from .serializers import SaleSerializer, SaleDetailSerializer
from .forms import SaleForm
from .utils import generate_invoice_pdf

class SaleViewSet(viewsets.ModelViewSet):
    """API ViewSet para ventas"""
    queryset = Sale.objects.all()
    serializer_class = SaleSerializer
    permission_classes = [permissions.IsAuthenticated]
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Filtrar por estado
        status = self.request.query_params.get('status')
        if status:
            queryset = queryset.filter(status=status)
        
        # Filtrar por fecha
        date_from = self.request.query_params.get('date_from')
        date_to = self.request.query_params.get('date_to')
        if date_from and date_to:
            queryset = queryset.filter(date__range=[date_from, date_to])
        
        return queryset
    
    def perform_create(self, serializer):
        with transaction.atomic():
            sale = serializer.save(created_by=self.request.user)
            
            # Registrar en auditoría
            AuditLog.objects.create(
                user=self.request.user,
                action='CREATE',
                model_name='Sale',
                record_id=sale.id,
                data_after=serializer.data,
                ip_address=self.request.META.get('REMOTE_ADDR'),
                user_agent=self.request.META.get('HTTP_USER_AGENT')
            )
            
            # Actualizar inventario
            for detail in sale.details.all():
                product = detail.product
                product.update_stock(detail.quantity, 'subtract')
                
                # Registrar movimiento de inventario
                InventoryMovement.objects.create(
                    product=product,
                    movement_type='SALE',
                    quantity=detail.quantity,
                    previous_stock=product.stock + detail.quantity,
                    new_stock=product.stock,
                    reference=sale.invoice_number,
                    created_by=self.request.user
                )
    
    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        """Cancelar una venta"""
        sale = self.get_object()
        if sale.status == 'CANCELLED':
            return Response(
                {'error': 'La venta ya está cancelada'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        with transaction.atomic():
            # Restaurar inventario
            for detail in sale.details.all():
                product = detail.product
                product.update_stock(detail.quantity, 'add')
                
                InventoryMovement.objects.create(
                    product=product,
                    movement_type='RETURN',
                    quantity=detail.quantity,
                    previous_stock=product.stock - detail.quantity,
                    new_stock=product.stock,
                    reference=f"CANCEL-{sale.invoice_number}",
                    notes=f"Cancelación de venta {sale.invoice_number}",
                    created_by=request.user
                )
            
            sale.status = 'CANCELLED'
            sale.save()
            
            AuditLog.objects.create(
                user=request.user,
                action='UPDATE',
                model_name='Sale',
                record_id=sale.id,
                data_after={'status': 'CANCELLED'},
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT')
            )
        
        return Response({'message': 'Venta cancelada exitosamente'})
    
    @action(detail=True, methods=['get'])
    def pdf(self, request, pk=None):
        """Generar PDF de factura"""
        sale = self.get_object()
        pdf_content = generate_invoice_pdf(sale)
        
        response = HttpResponse(pdf_content, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="factura_{sale.invoice_number}.pdf"'
        return response

@login_required
def sale_list(request):
    """Vista para listar ventas"""
    sales = Sale.objects.all().select_related('client', 'created_by')
    return render(request, 'sales/list.html', {'sales': sales})

@login_required
def sale_detail(request, pk):
    """Vista de detalle de venta"""
    sale = get_object_or_404(
        Sale.objects.select_related('client', 'created_by').prefetch_related('details__product'),
        pk=pk
    )
    return render(request, 'sales/detail.html', {'sale': sale})


@login_required
def sale_delete(request, pk):
    """Eliminar una venta y restaurar su inventario."""
    sale = get_object_or_404(Sale, pk=pk)
    if request.method == 'POST':
        with transaction.atomic():
            for detail in sale.details.all():
                product = detail.product
                product.update_stock(detail.quantity, 'add')
                InventoryMovement.objects.create(
                    product=product,
                    movement_type='RETURN',
                    quantity=detail.quantity,
                    previous_stock=product.stock - detail.quantity,
                    new_stock=product.stock,
                    reference=f"DELETE-{sale.invoice_number}",
                    notes=f"Eliminación de venta {sale.invoice_number}",
                    created_by=request.user,
                )
            sale.delete()
            AuditLog.objects.create(
                user=request.user,
                action='DELETE',
                model_name='Sale',
                record_id=pk,
                data_after={'deleted': True},
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
            )
        return redirect('sales:list')
    return redirect('sales:list')


def _get_sale_for_print(pk):
    return get_object_or_404(
        Sale.objects.select_related('client', 'created_by').prefetch_related('details__product'),
        pk=pk
    )


@login_required
def sale_receipt(request, pk):
    """Comprobante de venta listo para imprimir"""
    sale = _get_sale_for_print(pk)
    company = CompanyConfig.get_config()
    return render(request, 'sales/receipt.html', {
        'sale': sale,
        'company': company,
        'auto_print': request.GET.get('auto_print') == '1',
        'format': request.GET.get('format', 'ticket'),
    })


@login_required
def sale_print_pdf(request, pk):
    """Descargar factura en PDF"""
    sale = _get_sale_for_print(pk)
    pdf_content = generate_invoice_pdf(sale)

    AuditLog.objects.create(
        user=request.user,
        action='PRINT',
        model_name='Sale',
        record_id=sale.id,
        data_after={'invoice_number': sale.invoice_number, 'format': 'pdf'},
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
    )

    response = HttpResponse(pdf_content, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="factura_{sale.invoice_number}.pdf"'
    return response


@login_required
def sale_create(request):
    """Vista para crear una nueva venta"""
    form = SaleForm(request.POST or None)
    
    if request.method == 'POST' and form.is_valid():
        try:
            with transaction.atomic():
                sale = form.save(commit=False)
                sale.created_by = request.user
                sale.shift_number = int(request.session.get('sales_current_shift', 1))
                sale.subtotal = Decimal('0.00')
                sale.tax = Decimal('0.00')
                sale.total = Decimal('0.00')
                sale.save()
                
                # Procesar detalles
                products_data = json.loads(request.POST.get('products_data', '[]'))
                for item in products_data:
                    product = Product.objects.get(id=item['product_id'])
                    SaleDetail.objects.create(
                        sale=sale,
                        product=product,
                        quantity=item['quantity'],
                        price=item['price'],
                        discount=item.get('discount', 0)
                    )
                
                # Calcular totales y cerrar venta
                sale.calculate_totals()
                sale.status = Sale.Status.PAID
                sale.save()

                # Actualizar inventario y registrar movimientos
                for detail in sale.details.all():
                    product = detail.product
                    previous_stock = product.stock
                    product.update_stock(detail.quantity, 'subtract')
                    InventoryMovement.objects.create(
                        product=product,
                        movement_type='SALE',
                        quantity=detail.quantity,
                        previous_stock=previous_stock,
                        new_stock=product.stock,
                        reference=sale.invoice_number,
                        created_by=request.user,
                    )

                AuditLog.objects.create(
                    user=request.user,
                    action='CREATE',
                    model_name='Sale',
                    record_id=sale.id,
                    data_after={
                        'invoice_number': sale.invoice_number,
                        'total': float(sale.total),
                        'status': sale.status,
                    },
                    ip_address=request.META.get('REMOTE_ADDR'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
                )

                return JsonResponse({
                    'success': True,
                    'invoice_number': sale.invoice_number,
                    'redirect_url': reverse('sales:receipt', args=[sale.id]) + '?auto_print=1',
                })
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            }, status=400)
    
    clients = Client.objects.filter(is_active=True)
    products = Product.objects.filter(is_active=True, stock__gt=0)
    
    return render(request, 'sales/form.html', {
        'form': form,
        'clients': clients,
        'products': products
    })

@login_required
def shift_close(request):
    """Mostrar el resumen del turno actual sin reiniciarlo aún."""
    today = timezone.now().date()
    current_shift = int(request.session.get('sales_current_shift', 1))
    sales = Sale.objects.filter(date__date=today, status=Sale.Status.PAID, shift_number=current_shift).select_related('client', 'created_by')
    total_sales = sales.aggregate(total=Sum('total'))['total'] or Decimal('0.00')
    total_count = sales.count()
    cash_sales = sales.filter(payment_method=Sale.PaymentMethod.CASH)
    cash_total = cash_sales.aggregate(total=Sum('total'))['total'] or Decimal('0.00')

    context = {
        'sales': sales,
        'total_sales': total_sales,
        'total_count': total_count,
        'cash_total': cash_total,
        'date': today,
        'cashier': request.user.get_full_name() or request.user.username,
    }
    return render(request, 'sales/shift_close.html', context)


@login_required
@require_http_methods(["POST"])
def shift_close_print(request):
    """Reiniciar el turno cuando se imprime el cierre del día."""
    current_shift = int(request.session.get('sales_current_shift', 1))
    request.session['sales_current_shift'] = current_shift + 1
    request.session.modified = True

    AuditLog.objects.create(
        user=request.user,
        action='UPDATE',
        model_name='Shift',
        record_id=current_shift,
        data_after={'shift_closed': True, 'next_shift': current_shift + 1},
        ip_address=request.META.get('REMOTE_ADDR'),
        user_agent=request.META.get('HTTP_USER_AGENT', '')[:255],
    )

    return JsonResponse({'success': True, 'next_shift': current_shift + 1})

@login_required
def sale_report(request):
    """Vista para reportes de ventas"""
    return render(request, 'sales/reports.html')

@login_required
def export_sales_excel(request):
    """Exportar ventas a Excel"""
    # Obtener parámetros de filtro
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    # Filtrar ventas
    sales = Sale.objects.all()
    if date_from and date_to:
        sales = sales.filter(date__range=[date_from, date_to])
    
    # Crear libro de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ventas"
    
    # Encabezados
    headers = ['N° Factura', 'Cliente', 'Fecha', 'Subtotal', 'Descuento', 'Impuesto', 'Total', 'Estado']
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # Datos
    for row, sale in enumerate(sales, 2):
        ws.cell(row=row, column=1, value=sale.invoice_number)
        ws.cell(row=row, column=2, value=sale.client.name)
        ws.cell(row=row, column=3, value=sale.date.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=4, value=float(sale.subtotal))
        ws.cell(row=row, column=5, value=float(sale.discount))
        ws.cell(row=row, column=6, value=float(sale.tax))
        ws.cell(row=row, column=7, value=float(sale.total))
        ws.cell(row=row, column=8, value=sale.get_status_display())
    
    # Guardar en respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="reporte_ventas.xlsx"'
    wb.save(response)
    
    return response